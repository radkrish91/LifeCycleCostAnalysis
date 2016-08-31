from flask import Flask, request, jsonify, session, render_template, redirect, url_for
from flask.ext.mysql import MySQL
import json
from array import *
import math
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from scipy.stats import norm
from matplotlib import mlab
# get_ipython().magic(u'matplotlib inline')
# get_ipython().magic(u'precision 4')
plt.style.use('ggplot')

import scipy.stats as stats
#import seaborn as sns
import random
from openpyxl import Workbook
from openpyxl import load_workbook

app = Flask(__name__)
mysql = MySQL()
 
# MySQL configurations
app.config['MYSQL_DATABASE_USER'] = 'root'
app.config['MYSQL_DATABASE_PASSWORD'] = 'Rutgers123'
app.config['MYSQL_DATABASE_DB'] = 'wimdata'
app.config['MYSQL_DATABASE_HOST'] = '128.6.237.242'
app.config["MYSQL_DATABASE_PORT"] = 3306
#app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql+pymysql://root:Rutgers123@128.6.237.242/wimdata'
#mysql.init_app(app)

@app.route('/getHighway', methods=['POST'])
def getHighway():
	req = request.json
	sri = req['sri']
	startMP = req['startMP']
	endMP = req['endMP']
	conn = mysql.connect()
	cursor = conn.cursor()
	query = "SELECT LANES, AADT, HIGHWAY_NAME, HIGHWAY_ID, LAT1, LON1 FROM highways WHERE  SRI= '"+ sri + "'" + " AND START_MP = '" + format(float(startMP), '.6f') + "'" + " AND END_MP = '" + format(float(endMP), '.6f') + "'"
	cursor.execute(query)
	posts= {}
	posts['items'] = [dict(LANES=str(row[0]), AADT=str(row[1]), HIGHWAY_NAME=str(row[2]), HIGHWAY_ID=str(row[3]), LAT1=str(row[4]), LON1=str(row[5]) ) for row in cursor.fetchall()]
	for item in posts['items']:
		lanes = item['LANES']
		aadt = item['AADT']
		highway_name = item['HIGHWAY_NAME']
		highway_id = item['HIGHWAY_ID']
		latitude = item['LAT1']
		longitude = item['LON1']
	return jsonify({'lanes' : lanes, 'aadt' : aadt, 'highway_name' : highway_name, 'highway_id' : highway_id, 'latitude' : latitude, 'longitude' : longitude})

@app.route('/mapfiller', methods=['POST'])
def mapfiller():
	req = request.json
	sri = req['sri']
	conn = mysql.connect()
	cursor = conn.cursor()
	cursor.execute("SELECT GeoJSON, SLD_MP_ST, SLD_MP_END FROM links WHERE  SRI= '"+sri+"'")
	posts= {}
	posts['items'] = [dict(start_milepost=row[1], end_milepost=row[2], geometry=row[0]) for row in cursor.fetchall()]
	geojsonstring = '{"type": "FeatureCollection","features": ['
	for item in posts['items']:
		geojsonstring = str(geojsonstring) + '{"type":"Feature","properties": { "start_milepost": "' + str(item['start_milepost']) + '", "end_milepost": "' + str(item['end_milepost']) + '", "color": "#600080"},"geometry":' + str(item['geometry']) + '},'
	
	geojsonstring = geojsonstring[:-1] + ']}'
	
	with open('static/js/maps.json', 'w') as outfile:
		outfile.write(geojsonstring)
	
	return jsonify({'geojsonstring' : geojsonstring})
	
@app.route('/getBridges', methods=['POST'])
def getBridges():
	latitude = []
	longitude = []
	features = []
	location = []
	year_built = []
	length = []
	width = []
	req = request.json
	sri = req['sri']
	conn = mysql.connect()
	cursor = conn.cursor()
	cursor.execute("SELECT LAT, LON, FEATURES_DESC_006A, LOCATION_009, YEAR_BUILT_027, STRUCTURE_LEN_MT_049, DECK_WIDTH_MT_052 FROM unq_br WHERE LRS_INV_ROUTE_013A= '"+sri+"'")
	posts= {}
	posts['items'] = [dict(lat=row[0], long=row[1], feat=row[2], loca=row[3], year=row[4], len=row[5], wid=row[6]) for row in cursor.fetchall()]
	
	for item in posts['items']:
		latitude.append(str(item['lat']))
		longitude.append(str(item['long']))
		features.append(str(item['feat']))
		location.append(str(item['loca']))
		year_built.append(str(item['year']))
		length.append(str(item['len']))
		width.append(str(item['wid']))
	
	return jsonify({'latitude' : latitude, 'longitude' : longitude, 'features' : features, 'locations' : location, 'year_built' : year_built, 'length' : length, 'width' : width})
	
@app.route('/getMilePost', methods=['POST'])
def getMilePost():
	req = request.json
	sri = req['sri']
	startMP = array('f', [])
	endMP = array('f', [])
	conn = mysql.connect()
	cursor = conn.cursor()
	cursor.execute("SELECT SLD_MP_ST, SLD_MP_END FROM links WHERE  SRI= '"+sri+"'")
	posts= {}
	posts['items'] = [dict(startMP=row[0], endMP=row[1]) for row in cursor.fetchall()]
	#items = list(cursor.fetchall())
	#print posts['items']
	for item in posts['items']:
		startMP.append(item['startMP'])
		endMP.append(item['endMP'])
	
	return jsonify({'StartMilePost' : np.max(startMP), 'EndMilePost' : np.max(endMP)})

@app.route('/sriGetter', methods=['GET'])
def sriGetter():
	conn = mysql.connect()
	cursor = conn.cursor()
	cursor.execute("SELECT DISTINCT SRI FROM links WHERE sri REGEXP '__$' ORDER BY SRI")
	posts= {}
	posts['items'] = [dict(sri=row[0]) for row in cursor.fetchall()]
	return jsonify(posts)
	
@app.route('/probabilisticOutput', methods=['POST'])
def probabilisticOutput():
	wb = load_workbook('LCCA table_draft_V8_SimpleQueueModel_2.xlsx')
	# print wb.get_sheet_names()
	ws = wb['Probabilistic LCCA']

	# wb.save('sample_book.xlsx') 


	# In[39]:

	#Fix values for both alternatives
	#############################Input#############################
	runs = ws['C14'].value

	#Get text info such as project name
	Project_name = ws['B6'].value
	Alt_A_name = ws['C17'].value
	Alt_B_name = ws['C18'].value

	#get info of the structure
	length = ws['B10'].value
	width = ws['B11'].value

	analysis_period = ws['B19'].value
	ADT = ws['B29'].value
	truckpercent = ws['B30'].value
	vot_p = ws['C32'].value #Value of time (Passenger cars)
	vot_t = ws['E32'].value #Value of time (Trucks)

	############### User/Social Cost Inputs ###############

	#workzone
	wz_distance = length/5280 #change from feet to mile

	lanesopen_M_AltA = ws['C36'].value   #number of lanes open during maintenance - alternative A
	lanesopen_M_AltB = ws['E36'].value   #number of lanes open during maintenance - alternative B
	lanesopen_R_AltA = ws['C37'].value   #number of lanes open during rehabilitation - alternative A
	lanesopen_R_AltB = ws['E37'].value   #number of lanes open during maintenance - alternative B
	Sn_AltA = ws['C41'].value #Speed under normal condition, free flow speed - alternative A
	Sn_AltB = ws['E41'].value #Speed under normal condition, free flow speed - alternative B
	Sm_AltA = ws['C42'].value #Speed under workzone condition -maintenance - alternative A
	Sm_AltB = ws['E42'].value #Speed under workzone condition -maintenance - alternative B
	Sr_AltA = ws['C43'].value #Speed under workzone condition -rehab - alternative A
	Sr_AltB = ws['E43'].value #Speed under workzone condition -rehab - alternative B


	capacity_m_A = lanesopen_M_AltA*1480 #1480 may need to change to more detail number
	capacity_m_B = lanesopen_M_AltB*1480
	capacity_r_A = lanesopen_R_AltA*1480
	capacity_r_B = lanesopen_R_AltB*1480

	#VOC
	r = 8 # $8

	#Crash
	Ca = 142000
	Aa = 2.66
	An = 1.77

	#factor
	usercost_factor = ws['B25'].value # user cost weighted factor - show in project info tab
	socialcost_factor = ws['B26'].value # social cost weighted factor - show in project info tab

	#number of lanes
	n_inboundlanes = ws['C31'].value # lanes opened during normal condition - alternative A
	n_outboundlanes = ws['E31'].value # lanes opened during normal condition - alternative B

	#capacity non-workzone
	capacity_in = n_inboundlanes*1900
	capacity_out = n_outboundlanes*1900

	timeofday = range(0,24,1)
	dis_factor = [1.2, 0.8, 0.7, 0.5, 0.7, 1.7, 5.1, 7.8, 6.3, 5.2, 4.7, 5.3,
				  5.6, 5.7, 5.9, 6.5, 7.9, 8.5, 5.9, 3.9, 3.3, 2.8, 2.3, 1.7] #Distribution Factor

	inbound_factor = [47,43,46,48,57,58,63,60,59,55,46,49,
			   50,50,49,46,45,40,46,48,47,47,48,45]

	outbound_factor = [53,57,54,52,43,42,37,40,41,45,54,51,
				50,50,51,54,55,60,54,52,53,53,52,55]

	inbound_traffic =[]
	outbound_traffic = []

	for i in range(0,len(dis_factor)):
		inbound_traffic.append(round(ADT*dis_factor[i]/100*inbound_factor[i]/100,0))
		outbound_traffic.append(round(ADT*dis_factor[i]/100*outbound_factor[i]/100,0))

	print "------------------------------------"
	print "Alternative A:", Alt_A_name
	print "Alternative B:", Alt_B_name
	print "------------------------------------"


	# In[40]:

	# ############################# Probabilistic Input #############################
	rs = np.random.RandomState(5124)

	def select_dis(input_name, cell, par1, par2, par3):
		if cell == 'Normal':
			mu = par1
			sigma = par2
			input_name = rs.normal(mu, sigma, runs)
		elif cell == 'Lognormal':
	#         mu = par1
	#         sigma = par2
			mu = np.log(par1)
			sigma = np.log(par2)
	#         input_name = np.log(rs.lognormal(mu, sigma, runs))
			input_name = np.log(rs.lognormal(mu, sigma, runs))
		elif cell == 'Triangular':
			left = par1
			mode = par2
			right = par3
			input_name = rs.triangular(left, mode, right, runs)
		elif cell == 'Uniform':
			low = par1
			high = par2
			input_name = rs.uniform(low, high,runs)
		elif cell == 'Fixed':    
			value = par1
			input_name = np.repeat(value,runs)
		return input_name
			
					
	discountrate =[]
	discountrate = select_dis(discountrate, ws['D47'].value, ws['F47'].value, ws['H47'].value, ws['J47'].value)

	rehab_A = [] 
	rehab_B = []
	rehab_A = select_dis(rehab_A, ws['D48'].value, ws['F48'].value, ws['H48'].value, ws['J48'].value)
	rehab_B = select_dis(rehab_B, ws['D49'].value, ws['F49'].value, ws['H49'].value, ws['J49'].value)

	con_unitcost_A = []
	con_unitcost_B = []
	con_unitcost_A = select_dis(con_unitcost_A, ws['D50'].value, ws['F50'].value, ws['H50'].value, ws['J50'].value)
	con_unitcost_B = select_dis(con_unitcost_A, ws['D51'].value, ws['F51'].value, ws['H51'].value, ws['J51'].value)

	growthrate = []
	growthrate = select_dis(growthrate, ws['D52'].value, ws['F52'].value, ws['H52'].value, ws['J52'].value)

	maintenance_A = []
	maintenance_B = []
	maintenance_A = select_dis(maintenance_A, ws['D53'].value, ws['F53'].value, ws['H53'].value, ws['J53'].value)
	maintenance_B = select_dis(maintenance_B, ws['D54'].value, ws['F54'].value, ws['H54'].value, ws['J54'].value)

	duration_m_A = []
	duration_m_B = []
	duration_r_A = []
	duration_r_B = []
	duration_m_A = select_dis(duration_m_A, ws['D55'].value, ws['F55'].value, ws['H55'].value, ws['J55'].value)
	duration_m_B = select_dis(duration_m_B, ws['D56'].value, ws['F56'].value, ws['H56'].value, ws['J56'].value)
	duration_r_A = select_dis(duration_r_A, ws['D57'].value, ws['F57'].value, ws['H57'].value, ws['J57'].value)
	duration_r_B = select_dis(duration_r_B, ws['D58'].value, ws['F58'].value, ws['H58'].value, ws['J58'].value)

	#---------------------------------------------------------------------#
	# Calculate number of M&R activities to make it easier for loops later
	def num_activity(analysisperiod, schedule):
		num_activities = []
		for m in schedule:        
			if (analysisperiod%float(m)) ==0:
				temp = int(analysisperiod/float(m))-1
				num_activities.append(temp)
			else:
				temp = int(analysisperiod/float(m))
				num_activities.append(temp)
		return num_activities

	num_maintenance_A = num_activity(analysis_period, maintenance_A)
	num_rehab_A = num_activity(analysis_period, rehab_A)
	num_maintenance_B = num_activity(analysis_period, maintenance_B)
	num_rehab_B = num_activity(analysis_period, rehab_B)


	# In[41]:

	#############################
	#Agency Cost
	#############################
	print "Calculating Agency Cost..."

	############### Agency Cost Output #1: Initial Construction Cost ############### 
	initialcost_A = (con_unitcost_A*length*width)*(1+0.05)*(1+0.1+0.2)/1.8
	initialcost_B = (con_unitcost_B*length*width)*(1+0.05)*(1+0.1+0.2)/1.8

	print "Alt A: Initial construction cost($): Mean:", np.average(initialcost_A), "Std:", np.std(initialcost_A)
	print "Alt B: Initial construction cost($): Mean", np.average(initialcost_B), "Std:", np.std(initialcost_B)

	############### Agency Cost Output #2: Maintenance Cost ############### 

	percentage_m = 0.05 #Assume maintenance cost is 5% of the initial construction cost
	Cost_maintenance_A = []
	Cost_maintenance_B = []

	for j in range(len(maintenance_A)):
		Cost_m_temp_A = []
		for i in range(1,int(analysis_period/maintenance_A[j])):
			Cost_m_temp_A.append(percentage_m*initialcost_A[j]/((1+discountrate[j])**(i*maintenance_A[j])))
		Cost_maintenance_A.append(sum(Cost_m_temp_A))
	#print Cost_maintenance_C0
	print "Alt A: Maintenance cost($): Mean:",np.average(Cost_maintenance_A), "Std:", np.std(Cost_maintenance_A)

	for j in range(len(maintenance_B)):
		Cost_m_temp_B = []
		for i in range(1,int(analysis_period/maintenance_B[j])):
			Cost_m_temp_B.append(percentage_m*initialcost_B[j]/((1+discountrate[j])**(i*maintenance_B[j])))
		Cost_maintenance_B.append(sum(Cost_m_temp_B))
	#print Cost_maintenance_N0
	print "Alt B: Maintenance cost($): Mean:",np.average(Cost_maintenance_B), "Std:", np.std(Cost_maintenance_B)


	############### Agency Cost Output #3: Rehabilitation Cost ############### 

	Cost_rehab_A =[]
	Cost_rehab_B =[]

	for j in range(len(rehab_A)):
		Cost_rehab_temp_A =[]
		if (analysis_period/float(rehab_A[j])) <= 1:
			Cost_rehab_A.append(0)
		else:
			for i in range(1,int(analysis_period/rehab_A[j])+1):
				if ((i*rehab_A[j])%analysis_period) == 0:
					pass
				else:
					Cost_rehab_temp_A.append((initialcost_A[j]*1.8/(1+discountrate[j])**(i*rehab_A[j])))
			Cost_rehab_A.append(sum(Cost_rehab_temp_A))            
	print "Alt A: Rehabilitation cost($): Mean:",np.average(Cost_rehab_A), "Std:", np.std(Cost_rehab_A)

	for j in range(len(rehab_B)):
		Cost_rehab_temp_B =[]
		if (analysis_period/float(rehab_B[j])) <= 1:
			Cost_rehab_B.append(0)
		else:
			for i in range(1,int(analysis_period/rehab_B[j])+1):
				if ((i*rehab_B[j])%analysis_period) == 0:
					pass
				else:
					Cost_rehab_temp_B.append((initialcost_B[j]*1.8/(1+discountrate[j])**(i*rehab_B[j])))
			Cost_rehab_B.append(sum(Cost_rehab_temp_B))            
	print "Alt B: Rehabilitation cost($): Mean:",np.average(Cost_rehab_B), "Std:", np.std(Cost_rehab_B)

	############### Agency Cost Output #4: Salvage Value ###############

	salvage_A = []
	salvage_B = []

	for j in range(len(rehab_A)):
		temp_salvage_A = initialcost_A[j]*1.8*(((rehab_A[j]-analysis_period%rehab_A[j])/float(rehab_A[j]))/(1+discountrate[j])**analysis_period)
		salvage_A.append(temp_salvage_A)
	print "Alt A: Salvage value($): Mean:",-np.average(salvage_A), "Std:", -np.std(salvage_A)

	for j in range(len(rehab_B)):
		temp_salvage_B = initialcost_B[j]*1.8*(((rehab_B[j]-analysis_period%rehab_B[j])/float(rehab_B[j]))/(1+discountrate[j])**analysis_period)
		salvage_B.append(temp_salvage_B)
	print "Alt B: Salvage value($): Mean:",-np.average(salvage_B), "Std:", -np.std(salvage_B)


	############### Agency Cost Total ###############

	Agencycost_A = (initialcost_A + Cost_maintenance_A + Cost_rehab_A -salvage_A)/1000000
	Agencycost_B = (initialcost_B + Cost_maintenance_B +Cost_rehab_B - salvage_B)/1000000

	print "--------------------------------------------------------------------------------------"
	# print "Agency Cost for Alt A:", ("{0:.2f}".format(np.average(Agencycost_A))), "Million dollars"
	# print "Agency Cost for Alt B:",("{0:.2f}".format(np.average(Agencycost_B))), "Million dollars"

	######## Plot #######
	#sns.set(style="ticks")

	fig = plt.figure(figsize=(16,5))

	fig.add_subplot(121)
	(mu, sigma) = norm.fit(Agencycost_A)
	ws['C63'] = mu
	ws['E63'] = sigma
	print "Agency Cost for Alt A:", ("{0:.4f}".format(mu)), "Million dollars,", "std:",("{0:.4f}".format(sigma)),  "Million dollars"
	bins =80
	n, bins, patches = plt.hist(Agencycost_A, bins, normed=True, facecolor='blue', alpha=0.4,  label=("Alt A:"+Alt_A_name))
	y = mlab.normpdf(bins, mu, sigma)
	# l = plt.plot(bins, y, 'b--', linewidth=2)

	(mu2, sigma2) = norm.fit(Agencycost_B)
	ws['C64'] = mu2
	ws['E64'] = sigma2
	print "Agency Cost for Alt B:", ("{0:.4f}".format(mu2)), "Million dollars,", "std:",("{0:.4f}".format(sigma2)),  "Million dollars"
	print "--------------------------------------------------------------------------------------"
	bins2 =80
	n2, bins2, patches2 = plt.hist(Agencycost_B, bins2, normed=True, facecolor='indianred', alpha=0.4,  label=("Alt B:"+Alt_B_name))
	y2 = mlab.normpdf(bins2, mu2, sigma2)
	# l2 = plt.plot(bins2, y2, 'r--', linewidth=2)
	plt.legend(loc='best',fontsize=12)
	plt.ylabel("Probablity", fontsize=16)
	plt.xlabel("NPV (Million $)", fontsize=16)
	plt.title("Agency Cost PDF", fontsize=20)

	fig.add_subplot(122)
	#CDF
	n, bins, patches = plt.hist(Agencycost_A, bins, normed=True, histtype='step', cumulative=True, linewidth=1.2, color="blue",  label=("Alt A:"+Alt_A_name))
	y = mlab.normpdf(bins, mu, sigma).cumsum()
	y /= y[-1]
	# plt.plot(bins, y, 'k--', linewidth=1.5)
	n2, bins2, patches2 = plt.hist(Agencycost_B, bins2, normed=True, histtype='step', cumulative=True, linewidth=1.2, color="indianred", label=("Alt B:"+Alt_B_name))
	y2 = mlab.normpdf(bins2, mu2, sigma2).cumsum()
	y2 /= y2[-1]
	# plt.plot(bins2, y2, 'k--', linewidth=1.5)
	plt.legend(loc=2, fontsize=12)
	plt.ylabel("Cumulative Probablity", fontsize=16)
	plt.xlabel("NPV (Million $)", fontsize=16)
	plt.title("Agency Cost CDF", fontsize=20)
	plt.ylim(0.0,1.0)
	plt.tight_layout()
	plt.savefig("static/img/Agency_Cost.png")


	# In[42]:

	#############################
	#User Cost/Social Cost
	#############################
	print "Calculating User and Social Cost..."

	def MR_user_social_cost(num_MR, servicelife,normal_speed, wz_speed, wz_capacity, wz_duration, lanes_open):
		
		############### User Cost Output 1&2: Traffic Delay Cost (TDC) +Vehicle Operation Cost (V0C) ###############
		#Simple Queue Model
		allruns_MR_TDC = []
		allruns_MR_VOC = []
		all_stoppedveh_in = []
		all_stoppedveh_out = []
		for i in range(0,len(num_MR)):
			total_delaycost = []
			dis_stop_voc = []
			dis_idling_voc = []
			sum_stop_in = 0
			sum_stop_out = 0
			
			for nn in range(1,num_MR[i]+1):
				WZ_status = ['Y','Y','Y','Y','Y','Y','N','N','N','N','Y','Y','Y','Y','Y','N','N','N','N','Y','Y','Y','Y','Y']
				inb = []
				outb = []
				hour = []
				in_max = []
				out_max = []
				stoppedveh_in = 0
				stoppedveh_out = 0

				hour_counter = 0
				inb_sum_temp = 0
				out_sum_temp = 0

				
				for s in range(0, len(WZ_status)):
					if s == 0:
						hour_counter += 1
						inb_sum_temp += inbound_traffic[0]*(1+growthrate[i])**(servicelife[i]*nn)
						out_sum_temp += outbound_traffic[0]*(1+growthrate[i])**(servicelife[i]*nn)
						if WZ_status[s] == 'Y':
							in_max.append(capacity_r_A)
							out_max.append(capacity_r_A)
						else:
							in_max.append(capacity_in)
							out_max.append(capacity_out)
					else:
						if WZ_status[s] == WZ_status[s-1]:
							hour_counter += 1
							inb_sum_temp += inbound_traffic[s]*(1+growthrate[i])**(servicelife[i]*nn)
							out_sum_temp += outbound_traffic[s]*(1+growthrate[i])**(servicelife[i]*nn)
						else:
							inb.append(inb_sum_temp/hour_counter)
							outb.append(out_sum_temp/hour_counter)
							hour.append(hour_counter)
							if WZ_status[s] == 'Y':
								in_max.append(capacity_r_A)
								out_max.append(capacity_r_A)
							else:
								in_max.append(capacity_in)
								out_max.append(capacity_out)

							inb_sum_temp = inbound_traffic[s]*(1+growthrate[i])**(servicelife[i]*nn)
							out_sum_temp = outbound_traffic[s]*(1+growthrate[i])**(servicelife[i]*nn)
							hour_counter = 1
						if s == len(WZ_status)-1:
							inb.append(inb_sum_temp/hour_counter)
							outb.append(out_sum_temp/hour_counter)
							hour.append(hour_counter)

				max_q = 0 #A
				delay = []
				delay2 = []

				#inbound delay
				for s in range(0, len(inb)):
					if inb[s] > in_max[s]:
						stoppedveh_in += inb[s] * hour[s]
						max_q = hour[s]*(inb[s]-in_max[s])
						delay.append(max_q * hour[s] / 2)
					else:
						T = max_q/(in_max[s]-inb[s])
						delay.append(max_q * T / 2)
						stoppedveh_in += inb[s] * T
						max_q = 0

				if max_q > 0:
					inb_temp = sum(inbound_traffic[0:hour[0]])*(1+growthrate[i])**(servicelife[i]*nn)/hour[0]
					T = max_q/(in_max[0]-inb_temp)
					delay.append(max_q * T /2)
					stoppedveh_in += inb_temp * T

				#outbound delay
				max_q = 0       
				for s in range(0, len(outb)):
					if outb[s] > out_max[s]:
						stoppedveh_out += outb[s] * hour[s]
						max_q = hour[s]*(outb[s]-out_max[s])
						delay2.append(max_q * hour[s] / 2)
					else:
						T = max_q/(out_max[s]-outb[s])
						delay2.append(max_q * T / 2)
						stoppedveh_out += outb[s] * T
						max_q = 0

				if max_q > 0:
					out_temp = sum(outbound_traffic[0:hour[0]])*(1+growthrate[i])**(servicelife[i]*nn)/hour[0]
					T = max_q/(out_max[0]-out_temp)
					delay2.append(max_q * T /2) 
					stoppedveh_out += out_temp * T

				Total_delay_in = sum(delay)
				Total_delay_out = sum(delay2)
				Total_delay = Total_delay_in+Total_delay_out
				idling_voc = (Total_delay*(1-truckpercent)*0.6927+Total_delay*truckpercent*0.7681)*wz_duration[i]/2
				delaycost = Total_delay*((1-truckpercent)*vot_p+truckpercent*vot_t)*wz_duration[i]/2
				dis_delaycost = delaycost/(1+discountrate[i])**(servicelife[i]*nn)
				total_delaycost.append(dis_delaycost)
				sum_stop = stoppedveh_in + stoppedveh_out
				stop_voc = ((sum_stop)*(1-truckpercent)*109.02/1000+(sum_stop)*truckpercent*195.84/1000)*wz_duration[i]/2
				dis_stop_voc.append(stop_voc/((1+discountrate[i])**(servicelife[i]*nn)))
				dis_idling_voc.append(idling_voc/((1+discountrate[i])**(servicelife[i]*nn)))        
			allruns_MR_VOC.append(sum(dis_stop_voc)+sum(dis_idling_voc))
			allruns_MR_TDC.append(sum(total_delaycost))

		############### User Cost Output 3: Crash Risk Cost (CRC) ###############
		#Cost(Crash) = d*ADT*N*(Aa-An)*Ca
		#d=work zone distance(mile), ADT=average daily traffic(vpd), N=work zone durations(days)
		#Aa=crash rate under work zone conditions, An=crash rate under normal conditions, Ca=average crash cost ($/crash)

		allruns_MR_CRC = []
		for i in range(0,len(num_MR)):
			all_MR = []
			for n in range(1, num_MR[i]+1):
				future_traffic = ADT*(1+growthrate[i])**(n*servicelife[i])            
				a = wz_distance*(future_traffic)*wz_duration[i]*(Aa-An)*Ca/1000000
				MR_discounted = a/(1+discountrate[i])**(n*servicelife[i])
				all_MR.append(MR_discounted)
			allruns_MR_CRC.append(sum(all_MR))

		############### Social Cost Output 1: Air Pollution Cost (APC) ###############

		#Cost(air)=d*ADT*(0.01094+0.2155*F)*N
		#F=0.0723-0.00312(Vwz)+5.403*10^(-5)*(Vwz)^2
		#Vwz=Work zone speed(mph)

		allruns_MR_APC = []
		for i in range(0,len(num_MR)):
			all_MR = []
			for n in range(1, num_MR[i]+1):
				future_traffic = ADT*(1+growthrate[i])**(n*servicelife[i]) 
				a = wz_distance*(future_traffic)*(0.01094+0.2155*(0.0723-0.00312*(normal_speed)+5.403*10**(-5)*((normal_speed)**2)))*wz_duration[i]
				MR_discounted = a/(1+discountrate[i])**(n*servicelife[i])
				all_MR.append(MR_discounted)
			allruns_MR_APC.append(sum(all_MR))
		return allruns_MR_TDC, allruns_MR_VOC, allruns_MR_CRC, allruns_MR_APC


	TDC_r_A, VOC_r_A, CRC_r_A, APC_r_A = MR_user_social_cost(num_rehab_A, rehab_A, Sn_AltA, Sr_AltA, capacity_r_A, duration_r_A, lanesopen_R_AltA)
	TDC_r_B, VOC_r_B, CRC_r_B, APC_r_B = MR_user_social_cost(num_rehab_B, rehab_B, Sn_AltB, Sr_AltB, capacity_r_B, duration_r_B, lanesopen_R_AltB)

	# TDC_m_A, VOC_m_A, CRC_m_A, APC_m_A = MR_user_social_cost(num_maintenance_A, maintenance_A, Sn_AltA, Sm_AltA, capacity_m_A, duration_m_A, lanesopen_M_AltA)
	# TDC_m_B, VOC_m_B, CRC_m_B, APC_m_B = MR_user_social_cost(num_maintenance_B, maintenance_B, Sn_AltA, Sm_AltB, capacity_m_B, duration_m_B, lanesopen_M_AltB)

	# print TDC_r_A
	# print VOC_r_A
	# print CRC_r_A
	# print APC_r_A
	# print CRC_r_B
	# print APC_r_B

	usercost_A = []
	usercost_B = []
	socialcost_A = []
	socialcost_B = []

	for i in range(0,runs):
		usercost_A.append((TDC_r_A[i]+VOC_r_A[i]+CRC_r_A[i])/1000000)
		usercost_B.append((TDC_r_B[i]+VOC_r_B[i]+CRC_r_B[i])/1000000)
		socialcost_A.append((APC_r_A[i])/1000000)
		socialcost_B.append((APC_r_B[i])/1000000)

	LCC_A = []
	LCC_B = []

	for i in range(0,runs):    
		LCC_A.append(Agencycost_A[i]+usercost_factor*usercost_A[i]+socialcost_factor*socialcost_A[i])
		LCC_B.append(Agencycost_B[i]+usercost_factor*usercost_B[i]+socialcost_factor*socialcost_B[i])

	print "--------------------------------------------------------------------------------------"
	# print "LCC Alt A:", np.average(LCC_A)
	# print "LCC Alt B:", np.average(LCC_B)

	######## User Cost Plot #######
	#sns.set(style="ticks")

	fig = plt.figure(figsize=(16,5))

	fig.add_subplot(121)
	(mu, sigma) = norm.fit(usercost_A)
	ws['C67'] = mu
	ws['E67'] = sigma
	print "Average User Cost for Alt A:", ("{0:.4f}".format(mu)), "Million dollars,", "std:",("{0:.4f}".format(sigma)),  "Million dollars"
	bins =80
	n, bins, patches = plt.hist(usercost_A, bins, normed=True, facecolor='blue', alpha=0.4, label=("Alt A:"+Alt_A_name))
	y = mlab.normpdf(bins, mu, sigma)
	# l = plt.plot(bins, y, 'b--', linewidth=2)

	(mu2, sigma2) = norm.fit(usercost_B)
	ws['C68'] = mu2
	ws['E68'] = sigma2
	print "Average User Cost for Alt B:", ("{0:.4f}".format(mu2)), "Million dollars,", "std:",("{0:.4f}".format(sigma2)),  "Million dollars"
	print "--------------------------------------------------------------------------------------"
	bins2 =80
	n2, bins2, patches2 = plt.hist(usercost_B, bins2, normed=True, facecolor='indianred', alpha=0.4, label=("Alt B:"+Alt_B_name))
	y2 = mlab.normpdf(bins2, mu2, sigma2)
	# l2 = plt.plot(bins2, y2, 'r--', linewidth=2)
	plt.legend(loc='best',fontsize=14)
	plt.ylabel("Probablity", fontsize=16)
	plt.xlabel("NPV (Million $)", fontsize=16)
	plt.title("User Cost PDF", fontsize=20)

	fig.add_subplot(122)
	#CDF
	n, bins, patches = plt.hist(usercost_A, bins, normed=True, histtype='step', cumulative=True, linewidth=1.2, color= "blue",label=("Alt A:"+Alt_A_name))
	y = mlab.normpdf(bins, mu, sigma).cumsum()
	y /= y[-1]
	# plt.plot(bins, y, 'k--', linewidth=1.5)
	n2, bins2, patches2 = plt.hist(usercost_B, bins2, normed=True, histtype='step', cumulative=True, linewidth=1.2, color ="indianred", label=("Alt B:"+Alt_B_name))
	y2 = mlab.normpdf(bins2, mu2, sigma2).cumsum()
	y2 /= y2[-1]
	# plt.plot(bins2, y2, 'k--', linewidth=1.5)
	plt.legend(loc="best", fontsize=14)
	plt.ylabel("Cumulative Probablity", fontsize=16)
	plt.xlabel("NPV (Million $)", fontsize=16)
	plt.ylim(0.0,1.0)
	plt.title("User Cost CDF", fontsize=20)
	plt.tight_layout()
	plt.savefig("static/img/User_Cost.png")


	# In[43]:

	######## Social Cost Plot #######
	#sns.set(style="ticks")

	fig = plt.figure(figsize=(16,5))

	fig.add_subplot(121)
	(mu, sigma) = norm.fit(socialcost_A)
	ws['C71'] = mu
	ws['E71'] = sigma
	print "Average Social Cost for Alt A:", ("{0:.4f}".format(mu)), "Million dollars,", "std:",("{0:.4f}".format(sigma)),  "Million dollars"
	bins =80
	n, bins, patches = plt.hist(socialcost_A, bins, normed=True, facecolor='blue', alpha=0.4, label=("Alt A:"+Alt_A_name))
	y = mlab.normpdf(bins, mu, sigma)
	# l = plt.plot(bins, y, 'b--', linewidth=2)

	(mu2, sigma2) = norm.fit(socialcost_B)
	ws['C72'] = mu2
	ws['E72'] = sigma2
	print "Average Social Cost for Alt B:", ("{0:.4f}".format(mu2)), "Million dollars,", "std:",("{0:.4f}".format(sigma2)),  "Million dollars"
	print "--------------------------------------------------------------------------------------"
	bins2 =80
	n2, bins2, patches2 = plt.hist(socialcost_B, bins2, normed=True, facecolor='indianred', alpha=0.4, label=("Alt B:"+Alt_B_name))
	y2 = mlab.normpdf(bins2, mu2, sigma2)
	# l2 = plt.plot(bins2, y2, 'r--', linewidth=2)
	plt.legend(loc='best',fontsize=14)
	plt.ylabel("Probablity", fontsize=16)
	plt.xlabel("NPV (Million $)", fontsize=16)
	plt.title("Social Cost PDF", fontsize=20)

	fig.add_subplot(122)
	#CDF
	n, bins, patches = plt.hist(socialcost_A, bins, normed=True, histtype='step', cumulative=True, linewidth=1.2, color="blue", label=("Alt A:"+Alt_A_name))
	y = mlab.normpdf(bins, mu, sigma).cumsum()
	y /= y[-1]
	# plt.plot(bins, y, 'k--', linewidth=1.5)
	n2, bins2, patches2 = plt.hist(socialcost_B, bins2, normed=True, histtype='step', cumulative=True, linewidth=1.2, color="indianred",label=("Alt B:"+Alt_B_name))
	y2 = mlab.normpdf(bins2, mu2, sigma2).cumsum()
	y2 /= y2[-1]
	# plt.plot(bins2, y2, 'k--', linewidth=1.5)
	plt.legend(loc="best", fontsize=14)
	plt.ylabel("Cumulative Probablity", fontsize=16)
	plt.xlabel("NPV (Million $)", fontsize=16)
	plt.ylim(0.0,1.0)
	plt.title("Social Cost CDF", fontsize=20)
	plt.tight_layout()
	plt.savefig("static/img/Social_Cost.png")


	# In[44]:

	print "Calculating Total Life Cycle Cost..."


	######## LCCA Plot #######
	#sns.set(style="ticks")

	fig = plt.figure(figsize=(16,5))

	fig.add_subplot(121)
	(mu, sigma) = norm.fit(LCC_A)
	ws['C75'] = mu
	ws['E75'] = sigma
	print "Average NPV for Alt A:", ("{0:.4f}".format(mu)), "Million dollars,", "std:",("{0:.4f}".format(sigma)),  "Million dollars"
	bins =80
	n, bins, patches = plt.hist(LCC_A, bins, normed=True, facecolor='blue', alpha=0.4, label=("Alt A:"+Alt_A_name))
	y = mlab.normpdf(bins, mu, sigma)
	# l = plt.plot(bins, y, 'b--', linewidth=2)

	(mu2, sigma2) = norm.fit(LCC_B)
	ws['C76'] = mu2
	ws['E76'] = sigma2
	print "Average NPV for Alt B:", ("{0:.4f}".format(mu2)), "Million dollars,", "std:",("{0:.4f}".format(sigma2)),  "Million dollars"
	print "--------------------------------------------------------------------------------------"
	bins2 =80
	n2, bins2, patches2 = plt.hist(LCC_B, bins2, normed=True, facecolor='indianred', alpha=0.4, label=("Alt B:"+Alt_B_name))
	y2 = mlab.normpdf(bins2, mu2, sigma2)
	# l2 = plt.plot(bins2, y2, 'r--', linewidth=2)
	plt.legend(loc='best',fontsize=14)
	plt.ylabel("Probablity", fontsize=16)
	plt.xlabel("NPV (Million $)", fontsize=16)
	plt.title("Life Cycle Cost PDF", fontsize=20)

	fig.add_subplot(122)
	#CDF
	n, bins, patches = plt.hist(LCC_A, bins, normed=True, histtype='step', cumulative=True, linewidth=1.2, color="blue", label=("Alt A:"+Alt_A_name))
	y = mlab.normpdf(bins, mu, sigma).cumsum()
	y /= y[-1]
	plt.plot(bins, y, 'k--', linewidth=1.5)
	n2, bins2, patches2 = plt.hist(LCC_B, bins2, normed=True, histtype='step', cumulative=True, linewidth=1.2, color="indianred",label=("Alt B:"+Alt_B_name))
	y2 = mlab.normpdf(bins2, mu2, sigma2).cumsum()
	y2 /= y2[-1]
	plt.plot(bins2, y2, 'k--', linewidth=1.5)
	plt.legend(loc='best', fontsize=14)
	plt.ylabel("Cumulative Probablity", fontsize=16)
	plt.xlabel("NPV (Million $)", fontsize=16)
	plt.title("Life Cycle Cost CDF", fontsize=20)
	plt.tight_layout()
	plt.savefig("static/img/LCCA.png")
	wb.save('static/reports/Results.xlsx')
	#plt.show()
	return jsonify({'Status' : "success"})
	

@app.route('/output', methods=['POST'])
def output():
	wb = load_workbook('LCCA table_draft_V8_SimpleQueueModel_2.xlsx')
	ws = wb['Deterministic LCCA(hypo)']
	
	req = request.json
	print req
	_values = '[["Year", "Agency Cost(Conv mat)", "User Cost(Conv mat)", "Agency Cost(New mat)", "User Cost(New mat)",{ "role": "annotation" } ],'
	_pieValues = '[["Cost", "Total Amount"],'
	_pieValues2 = '[["Cost", "Total Amount"],'

	_length = req['length'] /5280
	_width = req['width']/ 5280
	_numberOfLanes = req['noLanes']
	_ADT = req['trafficVolume']
	_disc_rate = req['disc_rate']
	
	activityYears = [int(numeric_string) for numeric_string in req['activityYears']]
	activityNewYears = [int(numeric_string) for numeric_string in req['activityNewYears']]
	
	activityDays = [int(numeric_string) for numeric_string in req['activityDays']]
	activityNewDays = [int(numeric_string) for numeric_string in req['activityNewDays']]
	
	_years = array('i', activityYears);
	_new_years = array('i', activityNewYears)
	_days = array('i', activityDays);
	_new_days = array('i',activityNewDays)
	_analysis = int(req['analysis'])
	
	_replacementCostConv = 0;
	_resurfacingCostConv = 0;
	_totalMaintenanceCostConv = 0;
	_totalSewageCostConv = 0;
	_totalResurfacingCostConv = 0;
	_totalReplacementCostConv = 0;
	_replacementCostNew = 0;
	_resurfacingCostNew = 0;
	_totalTrafficDelayCostNew = 0;
	_totalcrashRiskCostNew = 0;
	_totalVehicleOperationCostNew = 0;
	_totalSewageCostNew = 0;
	_totalMaintenanceCostNew = 0;
	_totalResurfacingCostNew = 0;
	_totalReplacementCostNew = 0;
		
	_trafficSpeedDuringMaintenance = 40;
	_normalTrafficSpeed = 65;
	_hourlyTimeValue = 20;
	_weightedAverageVehicleCost = 8;
	
	_agencyCostConv = array('f', []);
	_userCostConv = array('f', []);
	
	_unitcostConv = 91.05;
	_unitcostNew = 106.50;
	_totalcrashRiskCostConv = 0;
	_totalTrafficDelayCostConv = 0;
	_totalVehicleOperationCostConv = 0;
	_price_factor = _unitcostNew / _unitcostConv;
	_initialConCostConv = 150 * _length *5280 * _width*5280;
	_agencyCostConv.append(_initialConCostConv); 
	_userCostConv.append(0);

	ws['B5'] = "projectSite"
	ws['B6'] = "constructionType"
	ws['B7'] = "state"
	ws['B8'] = "61.1901306"
	ws['D8'] = "61.3100552"
	ws['B9'] = req['length']
	ws['B10'] = req['width']
	ws['B11'] = "comments"
	ws['C14'] = "alternativeA"
	ws['C15'] = "alternativeB"
	ws['B16'] = _analysis
	ws['B17'] = _disc_rate
	ws['C18'] =  req['servlife']
	ws['E18'] = req['newservlife']
	ws['C19'] = req['cmat_val']
	ws['E19'] = req['nmat_val']
	ws['C20'] = req['ccon_val']
	ws['E20'] = req['ncon_val']
	ws['C21'] = req['conv_des']
	ws['E21'] = req['new_des']
	ws['C22'] = "0"
	ws['E22'] = "0"
	ws['B28'] = _ADT
	ws['B29'] = req['truck']
	ws['B30'] = "0.5"
	ws['C31'] = req['noLanes']
	ws['E31'] = req['noLanes']
	ws['C35'] = activityYears[0]
	ws['E35'] = activityNewYears[0]
	ws['C36'] = activityYears[2]
	ws['E36'] = activityNewYears[2]
	ws['C37'] = activityDays[0]
	ws['E37'] = activityNewDays[0]
	ws['C38'] = activityDays[2]
	ws['E38'] = activityNewDays[2]
	ws['C39'] = (_analysis/activityYears[0]) - 1
	ws['E39'] = (_analysis/activityNewYears[0]) - 1
	ws['C40'] = int(_analysis/activityDays[0])
	ws['E40'] = int(_analysis/activityNewDays[0])
	ws['C42'] = 2
	ws['E42'] = 2
	ws['C43'] = 1
	ws['E43'] = 1
	ws['C44'] = 1480
	ws['E44'] = 1900
	ws['B45'] = 15
	ws['C46'] = 65
	ws['E46'] = 65
	ws['C47'] = 40
	ws['E47'] = 40
	ws['C48'] = 20
	ws['E48'] = 20
	
	
	#for (_x = _years[0]; _x <= _analysis; _x=_x+_years[0])
	for x in range(_years[0], _analysis+1, _years[0]):
		discountedVal = pow((1+(_disc_rate/100)),x);
		_maintenanceCostConv = (0.05 * _initialConCostConv) / discountedVal ;
		if x == _years[1]:
			_vehicleOperationCostConv = (((_length / _trafficSpeedDuringMaintenance) - (_length / _normalTrafficSpeed)) * _ADT * _days[1] * _weightedAverageVehicleCost) / discountedVal;
			_trafficDelayCostConv = (((_length / _trafficSpeedDuringMaintenance) - (_length / _normalTrafficSpeed)) * _ADT * _days[1] * _hourlyTimeValue) / discountedVal;
			_crashRiskCostConv = (0.13 * _length * _ADT * _days[1]) / discountedVal;
			_resurfacingCostConv = (math.exp(-0.716 + (0.8197 * math.log(_numberOfLanes * _length))) * 1000000) / (discountedVal);
			_totalResurfacingCostConv = _totalResurfacingCostConv + _resurfacingCostConv;	
		elif x == _years[2]:
			_vehicleOperationCostConv = (((_length / _trafficSpeedDuringMaintenance) - (_length / _normalTrafficSpeed)) * _ADT * _days[2] * _weightedAverageVehicleCost) / discountedVal;
			_trafficDelayCostConv = (((_length / _trafficSpeedDuringMaintenance) - (_length / _normalTrafficSpeed)) * _ADT * _days[2] * _hourlyTimeValue) / discountedVal;
			_crashRiskCostConv = (0.13 * _length * _ADT * _days[2]) / discountedVal;
			_replacementCostConv = _initialConCostConv / discountedVal;
			_totalReplacementCostConv = _totalReplacementCostConv + _replacementCostConv;
		else:
			_vehicleOperationCostConv = (((_length / _trafficSpeedDuringMaintenance) - (_length / _normalTrafficSpeed)) * _ADT * _days[0] * _weightedAverageVehicleCost) / discountedVal;
			_crashRiskCostConv = (0.13 * _length * _ADT * _days[0]) / discountedVal;
			_trafficDelayCostConv = (((_length / _trafficSpeedDuringMaintenance) - (_length / _normalTrafficSpeed)) * _ADT * _days[0] * _hourlyTimeValue) / discountedVal;
		
		_agencyCostConv.append(_resurfacingCostConv + _replacementCostConv + _maintenanceCostConv);
		_userCostConv.append(_trafficDelayCostConv + _vehicleOperationCostConv + _crashRiskCostConv);
		
		_sewageCostConv = (0.02 * _initialConCostConv) / discountedVal;
		_totalTrafficDelayCostConv = _totalTrafficDelayCostConv + _trafficDelayCostConv;
		_totalcrashRiskCostConv = _totalcrashRiskCostConv + _crashRiskCostConv;
		_totalVehicleOperationCostConv = _totalVehicleOperationCostConv + _vehicleOperationCostConv;
		_totalMaintenanceCostConv = _totalMaintenanceCostConv + _maintenanceCostConv;
		_totalSewageCostConv = _totalSewageCostConv + _sewageCostConv;
		_resurfacingCostConv = 0;
		_replacementCostConv = 0;
		_maintenanceCostConv = 0;
		_trafficDelayCostConv = 0;
		_vehicleOperationCostConv = 0;
		_crashRiskCostConv = 0;

	_totalAgencyCostConv = _totalResurfacingCostConv + _totalReplacementCostConv + _totalMaintenanceCostConv + _initialConCostConv;		
	_totalUserCostConv = _totalTrafficDelayCostConv + _totalVehicleOperationCostConv + _totalcrashRiskCostConv;
	_lifeCycleCostConv = _totalAgencyCostConv + _totalUserCostConv;

	print "Agency cost of conventional material :%f" % _totalAgencyCostConv
	print "User cost of conventional material :%f" % _totalUserCostConv
	print "Life cycle cost of conventional material :%f" % _lifeCycleCostConv

	_agencyCostNew = array('f',[]);
	_userCostNew = array('f',[]);
	_initialConCostNew = 152.78 * _length *5280 * _width*5280;
	_agencyCostNew.append(_initialConCostNew);
	_userCostNew.append(0)
	
	for x in range(_new_years[0], _analysis+1, _new_years[0]):
		_discountedVal = pow((1+(_disc_rate/100)),x);
		_maintenanceCostNew = (0.05 * _initialConCostNew) / _discountedVal ;
		if x == _new_years[1]:
			_vehicleOperationCostNew = (((_length / _trafficSpeedDuringMaintenance) - (_length / _normalTrafficSpeed)) * _ADT * _new_days[1] * _weightedAverageVehicleCost) / _discountedVal;
			_trafficDelayCostNew = (((_length / _trafficSpeedDuringMaintenance) - (_length / _normalTrafficSpeed)) * _ADT * _new_days[1] * _hourlyTimeValue) / _discountedVal;
			_crashRiskCostNew = (0.13 * _length * _ADT * _new_days[1]) / _discountedVal;
			_resurfacingCostNew = ((math.exp(-0.716 + (0.8197 * math.log(_numberOfLanes * _length))) * 1000000) / (_discountedVal)) * _price_factor;
			_totalResurfacingCostNew = _totalResurfacingCostNew + _resurfacingCostNew;	
		elif x == _new_years[2]:
			_vehicleOperationCostNew = (((_length / _trafficSpeedDuringMaintenance) - (_length / _normalTrafficSpeed)) * _ADT * _new_days[2] * _weightedAverageVehicleCost) / _discountedVal;
			_trafficDelayCostNew = (((_length / _trafficSpeedDuringMaintenance) - (_length / _normalTrafficSpeed)) * _ADT * _new_days[2] * _hourlyTimeValue) / _discountedVal;
			_crashRiskCostNew = (0.13 * _length * _ADT * _new_days[2]) / _discountedVal;
			if _analysis <= _new_years[2]:
				_replacementCostNew = 0
			else:
				_replacementCostNew = _initialConCostNew / _discountedVal;
			_totalReplacementCostNew = _totalReplacementCostNew + _replacementCostNew;
		else:
			_vehicleOperationCostNew = (((_length / _trafficSpeedDuringMaintenance) - (_length / _normalTrafficSpeed)) * _ADT * _new_days[0] * _weightedAverageVehicleCost) / _discountedVal;
			_crashRiskCostNew = (0.13 * _length * _ADT * _new_days[0]) / _discountedVal;
			_trafficDelayCostNew = (((_length / _trafficSpeedDuringMaintenance) - (_length / _normalTrafficSpeed)) * _ADT * _new_days[0] * _hourlyTimeValue) / _discountedVal;

		_agencyCostNew.append(_resurfacingCostNew + _replacementCostNew + _maintenanceCostNew);
		_userCostNew.append(_trafficDelayCostNew + _vehicleOperationCostNew + _crashRiskCostNew);
		
		_sewageCostNew = (0.02 * _initialConCostNew) / _discountedVal;
		_totalTrafficDelayCostNew = _totalTrafficDelayCostNew + _trafficDelayCostNew;
		_totalcrashRiskCostNew = _totalcrashRiskCostNew + _crashRiskCostNew;
		_totalVehicleOperationCostNew = _totalVehicleOperationCostNew + _vehicleOperationCostNew;
		
		_totalSewageCostNew = _totalSewageCostNew + _sewageCostNew;
		_totalMaintenanceCostNew = _totalMaintenanceCostNew + _maintenanceCostNew;
		_totalSewageCostNew = _totalSewageCostNew + _sewageCostNew;
		_resurfacingCostNew = 0;
		_replacementCostNew = 0;
		_maintenanceCostNew = 0;
		_trafficDelayCostNew = 0;
		_vehicleOperationCostNew = 0;
		_crashRiskCostNew = 0;
	
	_totalAgencyCostNew = _totalResurfacingCostNew + _totalReplacementCostNew + _totalMaintenanceCostNew + _initialConCostNew;		
	_totalUserCostNew = _totalTrafficDelayCostNew + _totalVehicleOperationCostNew + _totalcrashRiskCostNew;
	_lifeCycleCostNew = _totalAgencyCostNew + _totalUserCostNew;

	print "Agency cost of new material :%f" % _totalAgencyCostNew
	print "User cost of new material :%f" % _totalUserCostNew
	print "Life cycle cost of new material :%f" % _lifeCycleCostNew

	i = 0;
	for x in range(0, _analysis+1, _years[0]):
		_valueToAdd = '["' + str(x) + '",' + `int(_agencyCostConv[i])` + ',' + `int(_userCostConv[i])` + ',' + `int(_agencyCostNew[i])` + ',' + `int(_userCostNew[i])` + ', ""],';
		_values = _values + _valueToAdd
		i = i + 1

	_values = _values[:-1] + ']'

	
	_pieValues = _pieValues + '["Initial Construction Cost", ' + `int(_initialConCostConv)` + '], ["Traffic Delay Cost", ' + `int(_totalTrafficDelayCostConv)` + '], ["Vehicle Operation Cost", ' + `int(_totalVehicleOperationCostConv)` + '], ["Crash Risk Cost", ' + `int(_totalcrashRiskCostConv)` + '], ["Resurfacing Cost", ' + `int(_totalResurfacingCostConv)` + '], ["Replacement Cost", ' + `int(_totalReplacementCostConv)` + '], ["Maintenance Cost", ' + `int(_totalMaintenanceCostConv)` + ']';
	_pieValues2 = _pieValues2 + '["Initial Construction Cost", ' + `int(_initialConCostNew)` + '], ["Traffic Delay Cost", ' + `int(_totalTrafficDelayCostNew)` + '], ["Vehicle Operation Cost", ' + `int(_totalVehicleOperationCostNew)` + '], ["Crash Risk Cost", ' + `int(_totalcrashRiskCostNew)` + '], ["Resurfacing Cost", ' + `int(_totalResurfacingCostNew)` + '], ["Replacement Cost", ' + `int(_totalReplacementCostNew)` + '], ["Maintenance Cost", ' + `int(_totalMaintenanceCostNew)` + ']';
	_pieValues = _pieValues + ']'
	_pieValues2 = _pieValues2 + ']'
	
	print _values
	print _pieValues
	print _pieValues2
	
	
	ws['C52'] = _initialConCostConv
	ws['E52'] = _initialConCostNew
	ws['C53'] = _totalMaintenanceCostConv
	ws['E53'] = _totalMaintenanceCostNew
	ws['C55'] = _totalReplacementCostConv
	ws['E55'] = _totalReplacementCostNew
	ws['C56'] = 20
	ws['E56'] = 20
	ws['C57'] = 20
	ws['E57'] = 20
	ws['C58'] = 20
	ws['E58'] = 20
	ws['C59'] = 20
	ws['E59'] = 20
	ws['C60'] = 20
	ws['E60'] = 20
	ws['C61'] = _totalAgencyCostConv
	ws['E61'] = _totalAgencyCostNew
	
	ws['C64'] = _totalTrafficDelayCostConv
	ws['E64'] = _totalTrafficDelayCostNew
	ws['C65'] = _totalVehicleOperationCostConv
	ws['E65'] = _totalVehicleOperationCostNew
	ws['C66'] = _totalcrashRiskCostConv
	ws['E66'] = _totalcrashRiskCostNew
	ws['C67'] = _totalUserCostConv
	ws['E67'] = _totalUserCostNew
	
	ws['C70'] = 20
	ws['E70'] = 20
	ws['C71'] = 20
	ws['E71'] = 20
	ws['C72'] = 20
	ws['E72'] = 20
	ws['C73'] = 20
	ws['E73'] = 20
	
	ws['C79'] = _lifeCycleCostConv
	ws['E79'] = _lifeCycleCostNew
	ws['B80'] = "alternative A"
	ws['D81'] = "20%"
	ws['D82'] = "20%"
	ws['D83'] = "20%"
	ws['E84'] = "20%"
	
	wb.save('static/reports/DeterministicResults.xlsx')
	
	return jsonify({'ColumnChartValues' : _values, 'PieChartValues' : _pieValues, 'PieChartValues2' : _pieValues2, 'lifeCycleCostConv' : _lifeCycleCostConv, 'lifeCycleCostNew' : _lifeCycleCostNew})
	
@app.route("/")
def main():
    return render_template('index.html')
if __name__ == "__main__":
	app.run()